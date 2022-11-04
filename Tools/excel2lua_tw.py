# -*- coding: utf-8 -*-
#!/usr/bin/python
import xlrd
import pycel
import openpyxl
import os
from sys import argv
from encrypt import encrypt
import check_syntax

def loadFilePaths(path = '文件路径.txt'):
    print("正在解析路径列表: ", path)
    paths = []
    with open(path, 'rb') as f:
        text = f.read()
        text = text.decode('utf-8')
        paths = eval(text)
        for path in paths:
            path[0] = path[0] + '.xlsx'
            path[2] = path[2] + '.lua'
            #print(path)

        return paths
    print("[Error]: 解析路径列表失败")
    return paths

def data2int(data):
    try:
        if type(data) == str and data != '':
            data = int(data)
    except Exception as e:
        print('[Error]: Index Error in data 2 int: ', e)
    finally:
        return data

# tag = 0, 处理普通数据; tag = 1处理id数据
def formatData(data, tag = 0):
    try:
        if data == None or data == '':
            data = 'nil'
        elif type(data) == float:
            if data.is_integer():
                data = int(data)
            else:
                data = round(data, 5)
        # elif tag == 0:
        #     data = '"' + data + '"'
        return data
    except Exception as e:
        print("[FormatData Error]: ", e, data)
        return 'nil'

def get_cell_value(cell, sheetName, excel):
    if cell.data_type == 'f':
        try:
            return excel.evaluate(f'{sheetName}!{cell.coordinate}')
        except Exception as e:
            raise Exception('Error in cell: ' + str(cell.coordinate) + ' Error: ' + str(e))
    else:
        return cell.value

def parse_excel(pathList):
    excelPath, sheetName, luaPath, encryptPath = pathList
    print('开始解析表: ', excelPath, '→', sheetName, '，目标路径: ',luaPath)
    try:
        book = openpyxl.load_workbook(excelPath)
    except Exception as e:
        print('[Error]: 找不到该excel文件: ', e)
        return

    # print("The number of worksheets is {0}".format(book.nsheets))
    # print("Worksheet name(s): {0}".format(book.sheet_names()))
    # 找到相应的Sheet
    sh = None
    for sn in book.sheetnames:
        if sn == sheetName:
            sh = book[sn]
            break

    if sh == None:
        print('[Error]: 没有找到对应的Sheet: ' + excelPath + '→' + sheetName + '，请检查配置文件')
        return None
    try:
        cn_excel = pycel.ExcelCompiler(excelPath)
    except Exception as e:
        print('[Error]: ExcelCompiler Error: ', e)
        return None
    booktw = None
    twexcelpath = excelPath.replace("导表excel", "导表excel翻译")
    if os.path.exists(twexcelpath) :
        booktw = openpyxl.load_workbook(twexcelpath)

    shtw = None
    if booktw:
        for sn in booktw.sheetnames:
            if sn == sheetName:
                shtw = booktw[sn]
        tw_excel = pycel.ExcelCompiler(twexcelpath)

    ##################
    os.makedirs(os.path.dirname(luaPath), exist_ok=True)
    # 开始写数据
    with open(luaPath, 'w', encoding='utf8') as targetFile:
        # targetFile = open(luaPath, 'w')
        # ---------- write -------------
        targetFile.write('local value_list = {} \n')
        # 从第四行开始是数据
        for rx in range(4, sh.max_row+1):
            # 处理是否导表
            # 1.导表 2.将id转为字符串 -1.不导表
            isload = get_cell_value(sh.cell(row=rx, column=1), sheetName, cn_excel)
            if isload == '' or isload == None or isload == -1 or isload == 0 or data2int(isload) == -1:
                continue
            id = get_cell_value(sh.cell(row=rx, column=2), sheetName, cn_excel)
            id = formatData(id, 1)
            if not id: continue
            if isload == 2:
                id = '"' + str(id) + '"'
            # print("id = ", type(id), id)
            targetFile.write('value_list[{0}] = '.format(id))
            targetFile.write('{')
            first_col = True
            # 从第二列开始是数据
            for cx in range(2, sh.max_column+1):
                # 空.导表 2. 导出为字符串 -1. 不导表
                key = get_cell_value(sh.cell(row=2, column=cx), sheetName, cn_excel)
                if not key: continue
                isload2 = get_cell_value(sh.cell(row=3, column=cx), sheetName, cn_excel)
                # print(sh.cell_value(rowx=1, colx=cx))
                if isload2 == -1 or data2int(isload2) == -1:
                    continue
                cell_value = get_cell_value(sh.cell(row=rx, column=cx), sheetName, cn_excel)
                cell_value = formatData(cell_value)
                if cell_value == None or cell_value == '' or cell_value == 'nil' or cell_value=='#REF!':
                    continue
                if isload2==2 or (isload==2 and key == 'id'):
                    cell_value = '"' + str(cell_value) + '"'
                # 处理title
                # title = sh.cell(row=1, column=cx).value
                # if title == '' or title == None:
                #     continue
                # 写 title = 数据,
                try:
                    if first_col:
                        first_col = False
                    else:
                        targetFile.write(', ')
                    targetFile.write(key)
                    targetFile.write(' = ')
                    targetFile.write(str(cell_value))
                except Exception as e:
                    print("[Error]: At ", rx, cx, e)
                    raise e
            targetFile.write('} \n')
        targetFile.write('\nreturn value_list\n')

    srctwPath = luaPath.replace("src", "srctw")
    twmap = {}
    twcxmap = {}
    if shtw:
        for rx in range(4, shtw.max_row+1):
            isload = get_cell_value(shtw.cell(row=rx, column=1), sheetName, tw_excel)
            if isload == '' or isload == None or isload == -1 or isload == 0 or data2int(isload) == -1:
                continue
            id = get_cell_value(shtw.cell(row=rx, column=2), sheetName, tw_excel)
            id = formatData(id, 1)
            if isload == 2:
                id = '"' + str(id) + '"'
            twmap[id] = rx
    
        for cx in range(2, shtw.max_column+1):
            isload = get_cell_value(shtw.cell(row=3, column=cx), sheetName, tw_excel)
            if isload == -1 or isload == 0 or data2int(isload) == -1:
                continue
            key = get_cell_value(shtw.cell(row=2, column=cx), sheetName, tw_excel)
            if not key: continue
            key = formatData(key)
            twcxmap[key] = cx

    # 开始写数据tw
    os.makedirs(os.path.dirname(srctwPath), exist_ok=True)
    with open(srctwPath, 'w', encoding='utf8') as targetFile:
        # targetFile = open(luaPath, 'w')
        # ---------- write -------------
        targetFile.write('local value_list = {} \n')
        # 从第三行开始是数据
        for rx in range(4, sh.max_row+1):
            # 处理是否导表
            # 1.导表 2.将id转为字符串 -1.不导表
            isload = get_cell_value(sh.cell(row=rx, column=1), sheetName, cn_excel)
            if isload == '' or isload == None or isload == -1 or isload == 0 or data2int(isload) == -1:
                continue
            id = get_cell_value(sh.cell(row=rx, column=2), sheetName, cn_excel)
            id = formatData(id, 1)
            if isload == 2:
                id = '"' + str(id) + '"'
            # print("id = ", type(id), id)
            targetFile.write('value_list[{0}] = '.format(id))
            targetFile.write('{')
            # 从第二列开始是数据
            # print("rx", rx, shtw.nrows)
            first_col = True
            for cx in range(2, sh.max_column+1):
                # 空.导表 2. 导出为字符串 -1. 不导表
                isload2 = get_cell_value(sh.cell(row=3, column=cx), sheetName, cn_excel)
                # print(sh.cell_value(rowx=1, colx=cx))
                if isload2 == -1 or data2int(isload2) == -1:
                    continue
                key = get_cell_value(sh.cell(row=2, column=cx), sheetName, cn_excel)
                if not key: continue
                key = formatData(key)
                cell_value = get_cell_value(sh.cell(row=rx, column=cx), sheetName, cn_excel)
                cell_value = formatData(cell_value)
                if cell_value == None or cell_value == '' or cell_value == 'nil':
                    continue
                if isload == 2 and cx == 2:
                    cell_value = '"' + str(cell_value) + '"'
                elif isload2 == 2:
                    if booktw and shtw:
                        tw_col = twcxmap[key]
                        if tw_col:
                            isloadtw2 = get_cell_value(shtw.cell(row=3, column=tw_col), sheetName, tw_excel)
                            if isloadtw2 == 3:
                                if twmap.get(id):
                                    cell_tw_tmp = shtw.cell(row=twmap[id], column=tw_col)
                                    cell_value_tw = get_cell_value(cell_tw_tmp, sheetName, tw_excel)
                                    cell_value_tw = formatData(cell_value_tw)
                                    if cell_value_tw == None or cell_value_tw == '' or cell_value_tw == 'nil':
                                        cell_value_tw = cell_value
                                    else:
                                        cell_value = cell_value_tw
                                    valuetw_id = get_cell_value(shtw.cell(row=twmap[id], column=2), sheetName, tw_excel)
                                    valuetw_id = formatData(valuetw_id, 1)
                                    if isload == 2:
                                        valuetw_id = '"' + str(valuetw_id) + '"'
                                else:
                                    print("[Error]: 导表excel翻译缺少id:", id)
                        else:
                            print("[Error]: 缺列", key)
                    cell_value = '"' + str(cell_value) + '"'
                # 处理title
                # title = sh.cell(row=2, column=cx).value
                # if title == '' or title == None:
                #     continue
                # 写 title = 数据,
                try:
                    if first_col:
                        first_col = False
                    else:
                        targetFile.write(', ')
                    targetFile.write(key)
                    targetFile.write(' = ')

                    targetFile.write(str(cell_value))
                    # if cx != sh.max_column:
                    #     targetFile.write(', ')
                except Exception as e:
                    print("[Error]: At ", rx, cx, e)

            targetFile.write('} \n')

        # --------- write end ----------
        # 合表
        # if luaPath.find('equipdata.lua') >= 0:
        #     targetFile.write('\nsetmetatable(value_list,{__index = require "data.item_data"})\n')
        # elif luaPath.find('item_data.lua') >= 0:
        #     targetFile.write('\nsetmetatable(value_list,{__index = require "data.component_data"})\n')
        # elif luaPath.find('gift_drop_data.lua') >= 0:
        #     targetFile.write('\nsetmetatable(value_list,{__index = require "data.drop_data"})\n')

        targetFile.write('\nreturn value_list\n')

    if shtw and shtw.max_column != sh.max_column:
        print("\n[Error]:导表excel与导表excel翻译行数不同！\n")
    if encryptPath:
        encrypt(luaPath, encryptPath)
        encrypt(srctwPath, encryptPath.replace("src", "srctw"))
        check_syntax.syntax(luaPath)
        check_syntax.syntax(srctwPath)

def parse_all_excels(paths):
    datas = []
    for path in paths:
        datas.append(parse_excel(path))
    return datas

def main():
    if len(argv) >= 2:
        paths = loadFilePaths(argv[1])
    else:
        paths = loadFilePaths()
    parse_all_excels(paths)

if __name__ == "__main__":
    print('>>>>>>>>>>>>>>>>>>>>>>')
    print('导表开始...')
    parse_excel(["..\\导表excel\\关卡\大富翁关卡\\【活动】导表_大富翁随机事件.xlsx", "随机事件", "..\src\\data\\monopoly_event_data.lua", None])
    print('导表结束...')
    print('<<<<<<<<<<<<<<<<<<<<<<')
    os.system("pause")
